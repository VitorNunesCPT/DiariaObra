#!/usr/bin/env python3

try:
    import pandas as pd
    import openpyxl
    from openpyxl import load_workbook
    
    # Carregar o arquivo Excel
    print("Analisando arquivo Excel...")
    wb = load_workbook('CONTROLE DE DIÁRIAS - FÁCIL.xlsx')
    
    print(f"\nPlanilhas encontradas:")
    for sheet_name in wb.sheetnames:
        print(f"- {sheet_name}")
    
    print(f"\nDetalhamento das planilhas:")
    for sheet_name in wb.sheetnames:
        print(f"\n=== {sheet_name} ===")
        ws = wb[sheet_name]
        
        # Verificar dimensões
        print(f"Dimensões: {ws.dimensions}")
        
        # Verificar primeiras linhas para entender estrutura
        print("Primeiras linhas (cabeçalhos):")
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
            if any(cell for cell in row if cell is not None):
                print(f"  {row}")
        
        # Verificar fórmulas e validações
        formulas = []
        validations = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    formulas.append(f"{cell.coordinate}: {cell.value}")
                if cell.data_validation:
                    validations.append(f"{cell.coordinate}: {cell.data_validation}")
        
        if formulas:
            print(f"Fórmulas encontradas: {len(formulas)}")
            for formula in formulas[:3]:  # Mostrar apenas as primeiras 3
                print(f"  {formula}")
        
        if validations:
            print(f"Validações encontradas: {len(validations)}")
            for validation in validations[:3]:  # Mostrar apenas as primeiras 3
                print(f"  {validation}")
    
    print(f"\n=== RESUMO DAS FUNCIONALIDADES IDENTIFICADAS ===")
    print("Baseado na análise do arquivo Excel, as funcionalidades incluem:")
    
    # Tentar identificar funcionalidades baseadas no conteúdo
    all_content = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            all_content.extend([str(cell) for cell in row if cell is not None])
    
    content_text = " ".join(all_content).lower()
    
    features = []
    if "data" in content_text or "dia" in content_text:
        features.append("Controle de datas")
    if "valor" in content_text or "preço" in content_text or "custo" in content_text:
        features.append("Controle de valores/gastos")
    if "categoria" in content_text or "tipo" in content_text:
        features.append("Categorização de gastos")
    if "descrição" in content_text or "obs" in content_text:
        features.append("Descrições/observações")
    if "total" in content_text or "soma" in content_text:
        features.append("Cálculos e totais")
    if "relatório" in content_text or "resumo" in content_text:
        features.append("Relatórios")
    if "filtro" in content_text or "busca" in content_text:
        features.append("Filtros")
    
    for feature in features:
        print(f"- {feature}")
    
    print(f"\nTotal de planilhas: {len(wb.sheetnames)}")
    print(f"Total de células com dados: {sum(1 for cell in all_content if cell.strip())}")

except ImportError as e:
    print(f"Erro: {e}")
    print("Tentando análise básica do arquivo...")
    
    # Análise básica sem bibliotecas
    try:
        with open('CONTROLE DE DIÁRIAS - FÁCIL.xlsx', 'rb') as f:
            content = f.read(1000)  # Ler primeiros 1000 bytes
            
        print("Análise básica do arquivo Excel:")
        print("- Arquivo Excel encontrado")
        print("- Tamanho: 87KB")
        print("- Provavelmente contém múltiplas planilhas")
        print("\nPara análise completa, seria necessário:")
        print("- pandas e openpyxl instalados")
        print("- ou acesso a ferramentas de análise de Excel")
        
    except Exception as e:
        print(f"Erro ao ler arquivo: {e}")

except Exception as e:
    print(f"Erro geral: {e}")